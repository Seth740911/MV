#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TMDB 电影刮削脚本
从 TMDB API 获取电影详细数据，补充到 Excel 中
图片保存到 I:\电影\.scraped\ 对应目录

配置说明：
- 如已开启 Clash TUN 模式：PROXIES = None（系统自动路由）
- 如未开启 TUN：PROXIES = {"http": "http://127.0.0.1:7897", ...}
"""

import os
import json
import re
import time
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ============================================================
# 配置区
# ============================================================
TMDB_API_KEY = "3109184759dc7c8960e1782250fd08d2"

# 代理配置：TUN 模式开启后设为 None 即可
# 未开启 TUN 时填入 Clash 混合端口（通常为 7897 或 7890）
# PROXIES = None  # TUN 模式已开启，无需显式代理
PROXIES = {"http": "http://127.0.0.1:7897", "https": "http://127.0.0.1:7897"}

TMDB_BASE = "https://api.themoviedb.org/3"
IMG_BASE = "https://image.tmdb.org/t/p/"
POSTER_SIZE = "w500"
BACKDROP_SIZE = "w1280"
PROFILE_SIZE = "w185"

EXCEL_PATH = r"G:\AI\MV\电影刮削数据.xlsx"
SCRAPED_DIR = r"I:\电影\.scraped"

# 每次请求间隔（秒），TMDB 免费版限速 40次/10秒
REQUEST_DELAY = 0.3

# ============================================================
# 带重试机制的 requests Session
# ============================================================

def create_session():
    """创建一个带自动重试的 requests Session"""
    session = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"]
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


SESSION = create_session()

# ============================================================
# 工具函数
# ============================================================

def tmdb_get(endpoint, params=None):
    """调用 TMDB API，自动重试"""
    if not TMDB_API_KEY:
        raise ValueError("请先填入 TMDB_API_KEY！")
    default = {"api_key": TMDB_API_KEY, "language": "zh-CN"}
    if params:
        default.update(params)
    resp = SESSION.get(
        f"{TMDB_BASE}{endpoint}",
        params=default,
        timeout=20,
        proxies=PROXIES,
    )
    resp.raise_for_status()
    return resp.json()


def tmdb_get_en(endpoint, params=None):
    """用英文获取（作为中文失败的备用）"""
    if not TMDB_API_KEY:
        return {}
    default = {"api_key": TMDB_API_KEY, "language": "en-US"}
    if params:
        default.update(params)
    try:
        resp = SESSION.get(
            f"{TMDB_BASE}{endpoint}",
            params=default,
            timeout=20,
            proxies=PROXIES,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception:
        return {}


def download_image(url, save_path):
    """下载图片到本地，带重试"""
    if not url:
        return False
    try:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
            return True
        resp = SESSION.get(url, timeout=30, proxies=PROXIES)
        resp.raise_for_status()
        with open(save_path, "wb") as f:
            f.write(resp.content)
        return True
    except Exception as e:
        print(f"  下载图片失败 {url}: {e}")
        return False


def search_movie(title, year=None):
    """搜索电影，返回 TMDB ID；中文无结果时自动用英文重试"""
    params = {"query": title, "page": 1}
    if year:
        params["year"] = year
    data = tmdb_get("/search/movie", params)
    results = data.get("results", [])
    if results:
        return results[0]["id"], results[0]
    # 中文无结果，换英文再试
    data_en = tmdb_get_en("/search/movie", params)
    results_en = data_en.get("results", [])
    if results_en:
        return results_en[0]["id"], results_en[0]
    return None, None


def get_movie_detail(tmdb_id):
    """获取电影详细信息（中文优先，英文备用）"""
    data_zh = tmdb_get(f"/movie/{tmdb_id}", {
        "append_to_response": "credits,keywords,release_dates"
    })
    data_en = tmdb_get_en(f"/movie/{tmdb_id}", {
        "append_to_response": "credits,keywords"
    })
    return data_zh, data_en


def get_directors(credits):
    """从 credits 中提取导演"""
    return [c["name"] for c in credits.get("crew", []) if c.get("job") == "Director"]


def get_cast(credits, limit=10):
    """获取主要演员姓名和 ID"""
    cast = credits.get("cast", [])[:limit]
    return [c["name"] for c in cast], [c["id"] for c in cast]


def get_countries(movie):
    """获取制片国家"""
    return [c["name"] for c in movie.get("production_countries", [])]


def get_genres(movie):
    """获取类型（TMDB 字段名为 genreres）"""
    return [g["name"] for g in movie.get("genres", [])]


def get_runtime(movie):
    """获取时长（分钟）"""
    return movie.get("runtime", None)


def get_overview(movie_zh, movie_en):
    """获取简介，中文优先；中文为空或太短时 fallback 到英文"""
    overview = movie_zh.get("overview", "")
    if not overview or len(overview) < 10:
        overview = movie_en.get("overview", "")
    return overview


def get_imdb_id(movie):
    """获取 IMDB ID"""
    return movie.get("imdb_id", "")


def download_poster(movie, tmdb_id):
    """下载海报，返回相对路径"""
    poster_path = movie.get("poster_path", "")
    if not poster_path:
        return ""
    ext = os.path.splitext(poster_path)[1] or ".jpg"
    filename = f"{tmdb_id}{ext}"
    save_path = os.path.join(SCRAPED_DIR, "posters", filename)
    url = f"{IMG_BASE}{POSTER_SIZE}{poster_path}"
    if download_image(url, save_path):
        return os.path.join(".scraped", "posters", filename).replace("\\", "/")
    return ""


def download_fanart(movie, tmdb_id):
    """下载背景图，返回相对路径"""
    backdrop_path = movie.get("backdrop_path", "")
    if not backdrop_path:
        return ""
    ext = os.path.splitext(backdrop_path)[1] or ".jpg"
    filename = f"{tmdb_id}{ext}"
    save_path = os.path.join(SCRAPED_DIR, "fanart", filename)
    url = f"{IMG_BASE}{BACKDROP_SIZE}{backdrop_path}"
    if download_image(url, save_path):
        return os.path.join(".scraped", "fanart", filename).replace("\\", "/")
    return ""


def download_stills(tmdb_id, limit=3):
    """下载剧照，返回相对路径列表"""
    try:
        data = tmdb_get(f"/movie/{tmdb_id}/images")
        stills = data.get("backdrops", [])[:limit]
        paths = []
        for i, s in enumerate(stills):
            file_path = s.get("file_path", "")
            if not file_path:
                continue
            ext = os.path.splitext(file_path)[1] or ".jpg"
            filename = f"{tmdb_id}_still{i}{ext}"
            save_path = os.path.join(SCRAPED_DIR, "stills", filename)
            url = f"{IMG_BASE}{BACKDROP_SIZE}{file_path}"
            if download_image(url, save_path):
                rel = os.path.join(".scraped", "stills", filename).replace("\\", "/")
                paths.append(rel)
        return paths
    except Exception as e:
        print(f"  获取剧照失败: {e}")
        return []


def download_actor_photos(cast_data, actor_img_map):
    """下载演员照片，更新 actor_img_map；返回新下载的演员名列表"""
    new_photos = []
    for name, pid in cast_data:
        if not name or name in actor_img_map:
            continue
        try:
            person = tmdb_get(f"/person/{pid}", {"language": "en-US"})
            profile_path = person.get("profile_path", "")
            if not profile_path:
                continue
            safe_name = "".join(c for c in name if c not in r'<>:"/\|?*')
            ext = os.path.splitext(profile_path)[1] or ".jpg"
            filename = f"{safe_name}{ext}"
            save_path = os.path.join(SCRAPED_DIR, "actors", filename)
            url = f"{IMG_BASE}{PROFILE_SIZE}{profile_path}"
            if download_image(url, save_path):
                rel_path = os.path.join(".scraped", "actors", filename).replace("\\", "/")
                actor_img_map[name] = rel_path
                new_photos.append(name)
                print(f"  演员照片: {name}")
        except Exception as e:
            print(f"  演员 [{name}] 照片下载失败: {e}")
        time.sleep(REQUEST_DELAY)
    return new_photos


# ============================================================
# 主流程
# ============================================================

def main():
    if not TMDB_API_KEY:
        print("=" * 60)
        print("请先填入 TMDB_API_KEY！")
        print("=" * 60)
        return

    # 加载 Excel
    print(f"读取 Excel: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    # 优先用 "电影" sheet，兼容旧版 "视频文件"
    if "电影" in wb.sheetnames:
        ws = wb["电影"]
    elif "视频文件" in wb.sheetnames:
        ws = wb["视频文件"]
    else:
        print("找不到「电影」或「视频文件」sheet！")
        print(f"可用 sheet: {wb.sheetnames}")
        return

    # 读取表头映射
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[cell.value] = cell.column

    print(f"表头列: {list(header_map.keys())}")

    # 确保需要的列存在，不存在则创建
    needed_cols = [
        "导演", "时长", "简介", "IMDB",
        "类型", "国家", "评分", "英文名",
        "TMDB_ID", "海报", "背景图", "剧照",
    ]
    max_col = ws.max_column
    for col_name in needed_cols:
        if col_name not in header_map:
            max_col += 1
            ws.cell(row=1, column=max_col, value=col_name)
            header_map[col_name] = ws.cell(row=1, column=max_col).column
            print(f"  新建列: {col_name} -> 列{ws.cell(row=1, column=max_col).column_letter}")

    # 重新构建表头映射（新建列后 column 属性可能变化）
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[cell.value] = cell.column

    # 加载已有 metadata.json 和 actor_img_map.json
    metadata_path = os.path.join(SCRAPED_DIR, "metadata.json")
    actor_map_path = os.path.join(SCRAPED_DIR, "actor_img_map.json")
    metadata = {}
    actor_img_map = {}
    if os.path.exists(metadata_path):
        with open(metadata_path, "r", encoding="utf-8") as f:
            metadata = json.load(f)
        print(f"已加载 metadata.json: {len(metadata)} 条")
    if os.path.exists(actor_map_path):
        with open(actor_map_path, "r", encoding="utf-8") as f:
            actor_img_map.update(json.load(f))
        print(f"已加载 actor_img_map.json: {len(actor_img_map)} 条")

    total = ws.max_row
    print(f"\n开始刮削，共 {total - 1} 部电影...\n")
    print(f"代理设置: {PROXIES}")
    print(f"请求间隔: {REQUEST_DELAY}s\n")

    updated = 0
    skipped = 0
    errors = 0

    for row in range(2, total + 1):
        name_cell = ws.cell(row=row, column=header_map.get("中文名", header_map.get("名称", 1)))
        name = name_cell.value
        year_val = ws.cell(row=row, column=header_map.get("年份", 2)).value
        imdb_val = ws.cell(row=row, column=header_map.get("IMDB", 1)).value

        # 提取年份数字
        year = None
        if year_val:
            m = re.search(r"(\d{4})", str(year_val))
            if m:
                year = int(m.group(1))

        # 跳过已有导演数据的行
        director_col = header_map.get("导演", 1)
        director_val = ws.cell(row=row, column=director_col).value
        if director_val and str(director_val).strip():
            print(f"[{row - 1}/{total - 1}] {name} ({year or '?'}) -> 已有导演，跳过")
            skipped += 1
            continue

        print(f"[{row - 1}/{total - 1}] {name} ({year or '?'}) ... ", end="", flush=True)

        # 尝试通过 IMDB ID 直接获取 TMDB ID
        tmdb_id = None
        if imdb_val and str(imdb_val).strip():
            try:
                find_data = tmdb_get("/find/" + str(imdb_val).strip(), {
                    "external_source": "imdb_id"
                })
                movie_results = find_data.get("movie_results", [])
                if movie_results:
                    tmdb_id = movie_results[0]["id"]
                    print(f"IMDB匹配 tmdb_id={tmdb_id} ... ", end="", flush=True)
            except Exception:
                pass

        # 通过名称和年份搜索
        if not tmdb_id:
            tmdb_id, _ = search_movie(name, year)
            if not tmdb_id and name:
                # 尝试去掉中文括号内容再搜
                clean_name = re.sub(r"[（(].*?[）)]", "", str(name)).strip()
                if clean_name != str(name):
                    tmdb_id, _ = search_movie(clean_name, year)

        if not tmdb_id:
            print("未找到匹配")
            skipped += 1
            time.sleep(REQUEST_DELAY)
            continue

        # 获取详细信息
        try:
            movie_zh, movie_en = get_movie_detail(tmdb_id)
        except Exception as e:
            print(f"获取详情失败: {e}")
            errors += 1
            time.sleep(REQUEST_DELAY)
            continue

        credits = movie_zh.get("credits", movie_en.get("credits", {}))

        # 导演
        directors = get_directors(credits)
        if directors:
            ws.cell(row=row, column=header_map.get("导演"), value=" / ".join(directors))

        # 时长
        runtime = get_runtime(movie_zh) or get_runtime(movie_en)
        if runtime:
            ws.cell(row=row, column=header_map.get("时长"), value=runtime)

        # 简介
        overview = get_overview(movie_zh, movie_en)
        if overview:
            cell = ws.cell(row=row, column=header_map.get("简介"), value=overview)
            cell.alignment = Alignment(wrap_text=True)

        # IMDB ID
        imdb_id = get_imdb_id(movie_zh) or get_imdb_id(movie_en)
        if imdb_id:
            ws.cell(row=row, column=header_map.get("IMDB"), value=imdb_id)

        # 类型
        genres = get_genres(movie_zh) or get_genres(movie_en)
        if genres:
            ws.cell(row=row, column=header_map.get("类型"), value=" / ".join(genres))

        # 国家
        countries = get_countries(movie_zh) or get_countries(movie_en)
        if countries:
            ws.cell(row=row, column=header_map.get("国家"), value=" / ".join(countries))

        # 评分
        rating = movie_zh.get("vote_average") or movie_en.get("vote_average")
        if rating:
            ws.cell(row=row, column=header_map.get("评分"), value=round(rating, 1))

        # 英文名
        title_en = movie_en.get("title", "") or movie_zh.get("original_title", "")
        if title_en:
            ws.cell(row=row, column=header_map.get("英文名"), value=title_en)

        # TMDB ID
        ws.cell(row=row, column=header_map.get("TMDB_ID"), value=tmdb_id)

        # 下载海报
        poster_rel = download_poster(movie_zh, tmdb_id)
        if poster_rel:
            ws.cell(row=row, column=header_map.get("海报"), value=poster_rel)

        # 下载背景图
        fanart_rel = download_fanart(movie_zh, tmdb_id)
        if fanart_rel:
            ws.cell(row=row, column=header_map.get("背景图"), value=fanart_rel)

        # 下载剧照
        stills = download_stills(tmdb_id, limit=3)
        if stills:
            ws.cell(row=row, column=header_map.get("剧照"), value=" | ".join(stills))

        # 下载演员照片
        cast_names, cast_ids = get_cast(credits, limit=10)
        cast_data = list(zip(cast_names, cast_ids))
        new_actors = download_actor_photos(cast_data, actor_img_map)

        # 更新 metadata.json
        metadata[str(tmdb_id)] = {
            "tmdb_id": tmdb_id,
            "imdb_id": imdb_id,
            "title_zh": movie_zh.get("title", ""),
            "title_en": title_en,
            "year": movie_zh.get("release_date", "")[:4] if movie_zh.get("release_date") else "",
            "director": " / ".join(directors),
            "cast": cast_names,
            "genres": genres,
            "countries": countries,
            "runtime": runtime,
            "rating": round(rating, 1) if rating else None,
            "overview": overview,
            "poster": poster_rel,
            "fanart": fanart_rel,
            "stills": stills,
            "has_poster": bool(poster_rel),
        }

        rating_str = round(rating, 1) if rating else "?"
        genre_str = " / ".join(genres[:2]) if genres else "?"
        print(f"导演:{directors[0] if directors else '?'} 评分:{rating_str} 类型:{genre_str}")
        updated += 1

        # 每 10 部保存一次（更频繁，防丢进度）
        if updated % 10 == 0:
            wb.save(EXCEL_PATH)
            with open(metadata_path, "w", encoding="utf-8") as f:
                json.dump(metadata, f, ensure_ascii=False, indent=2)
            with open(actor_map_path, "w", encoding="utf-8") as f:
                json.dump(actor_img_map, f, ensure_ascii=False, indent=2)
            print(f"  >> 已保存进度 ({updated} 部更新, {skipped} 部跳过, {errors} 部错误)\n")

        time.sleep(REQUEST_DELAY)

    # 最终保存
    print(f"\n保存 Excel...")
    wb.save(EXCEL_PATH)
    with open(metadata_path, "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)
    with open(actor_map_path, "w", encoding="utf-8") as f:
        json.dump(actor_img_map, f, ensure_ascii=False, indent=2)

    print(f"\n{'=' * 60}")
    print(f"刮削完成！")
    print(f"  更新: {updated} 部")
    print(f"  跳过: {skipped} 部（已有数据）")
    print(f"  错误: {errors} 部")
    print(f"  metadata.json: {len(metadata)} 条")
    print(f"  actor_img_map.json: {len(actor_img_map)} 条")
    print(f"  Excel 已保存: {EXCEL_PATH}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
